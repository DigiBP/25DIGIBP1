import time
import traceback
from art import art
from openpyxl import load_workbook
from pathlib import Path

from SupportFunctions import *


TOPIC = "withdraw_feedback"
WORKER_ID = "python-worker-16"



def set_status(business_key):
    """
    Update the status of a feedback entry in the Excel file to 'withdrawn'
    based on the provided business key.

    Args:
        business_key: Unique identifier of the feedback entry to update.
    """

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # get row with the business key
    row_number = None

    #search for row with process instance that should be changed
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[0]  # business key column
        if int(cell.value) == int(business_key):
            row_number = cell.row
            break

    # write to cell
    ws.cell(row_number,16, "withdrawn")

    # save excel
    wb.save(EXCEL_FILE)
    print("Data written to Excel.")




if __name__ == "__main__":
    print(f"Worker \"{Path(__file__).name}\" started — polling Camunda...")
    try:
        while True:
            try:
                for task in fetch_and_lock(worker_id=WORKER_ID, topic=TOPIC):
                    business_key = task.get("businessKey", "")
                    task_id = task["id"]
                    variables = {k: v["value"] for k, v in task["variables"].items()}
                    print(f"Worker \"{Path(__file__).name} fetched task {task_id} with business key {business_key}")
                    try:
                        set_status(business_key=business_key)
                        complete_task(task_id=task_id, variables=variables, worker_id=WORKER_ID)
                        print(f"Worker \"{Path(__file__).name} completed task {task_id} with business key {business_key}")
                    except PermissionError:
                        print("Excel in use, try again later")
                        time.sleep(30)
                    except Exception as exc:
                        print(f"Error in task {task_id}: {exc} {art('confused scratch')}")
                        traceback.print_exc()

            except Exception as exc:
                print(f"Fetch error: {exc} {art('table flip2')}")
                traceback.print_exc()

            time.sleep(5)

    except KeyboardInterrupt:
        pass