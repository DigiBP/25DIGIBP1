import time
import traceback
from pathlib import Path
from shutil import copyfile
from art import art
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, range_boundaries
from os.path import exists

from SupportFunctions import *



TOPIC     = "store_feedback_in_db"
WORKER_ID = "python-worker-17"

TEMPLATE_FILE = Path(__file__).parent / "feedback_db_template.xlsm"
BOLD          = Font(bold=True)


TABLE_NAME = "tblFeedback"
KEY_COL    = 1


def next_data_row(ws) -> int:
    """
    Returns the first row *after* the last row that has data in KEY_COL.
    Ignores rows that are only styled by the table.
    """
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=KEY_COL).value not in (None, ""):
            return r + 1
    return 2                      # sheet is empty except header


def get_or_create_workbook() -> "openpyxl.workbook.Workbook":
    """
    Returns an openpyxl workbook, always with keep_vba=True so macros survive.
    If the file does not exist it copies the TEMPLATE_FILE first.
    """
    db_path = Path(EXCEL_FILE)

    if db_path.exists():
        return load_workbook(db_path, keep_vba=True)

    # ---------- first-time initialisation ----------
    if TEMPLATE_FILE.exists():
        copyfile(TEMPLATE_FILE, db_path)
        print(f"Created new DB from template → {db_path.name}")
        return load_workbook(db_path, keep_vba=True)

    # Fallback: build a blank workbook (no VBA project)
    print("Template file not found – creating a blank .xlsm without macros.")
    wb = Workbook()
    ws = wb.active

    headers = [
        "businessKey",
        "feedbackDate",
        "feedbackType",
        "query",
        "email",
        "phone",
        "firstName",
        "lastName",
        "feedbackText",
        "needsClarification",
        "urgency",
        "impactScope",
        "forwardToDepartment",
        "linkToAdditForm",
        "reminderSent",
        "status",
        "measuresTaken"
    ]

    for col, name in enumerate(headers, start=1):
        ws.cell(1, col, name).font = BOLD

    wb.save(db_path)
    return wb


def write_to_excel(data: dict, business_key: str) -> None:
    wb = get_or_create_workbook()
    ws = wb.active

    row = next_data_row(ws)
    feedback_date = date.today().strftime("%d.%m.%Y")

    # ---- write cells ----
    ws.cell(row, 1,  business_key)
    ws.cell(row, 2,  feedback_date)
    ws.cell(row, 5,  data.get("email"))
    ws.cell(row, 6,  data.get("phone"))
    ws.cell(row, 7,  data.get("firstName"))
    ws.cell(row, 8,  data.get("lastName"))
    ws.cell(row, 9,  data.get("feedbackText"))
    ws.cell(row, 16, "open")

    # ---- if we outran the table, grow it ----
    if TABLE_NAME in ws.tables:
        tbl = ws.tables[TABLE_NAME]
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        if row > max_row:
            tbl.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{row}"
            )

    wb.save(EXCEL_FILE)
    print(f"Data written on row {row}.")



if __name__ == "__main__":
    print(f"Worker \"{Path(__file__).name}\" started — polling Camunda...")
    try:
        time.sleep(2)
        while True:
            try:
                for task in fetch_and_lock(worker_id=WORKER_ID, topic=TOPIC):
                    business_key = task.get("businessKey", "")
                    task_id = task["id"]
                    variables = {k: v["value"] for k, v in task["variables"].items()}
                    print(f"Worker \"{Path(__file__).name} fetched task {task_id} with business key {business_key}")
                    try:
                        write_to_excel(data=variables, business_key=business_key)
                        complete_task(task_id=task_id, variables=variables, worker_id=WORKER_ID)
                        print(f"Worker \"{Path(__file__).name} completed task {task_id} with business key {business_key}")
                    except Exception as exc:
                        print(f"Error in task {task_id}: {exc} {art('confused scratch')}")
                        traceback.print_exc()

            except Exception as exc:
                print(f"Fetch error: {exc} {art('table flip2')}")
                traceback.print_exc()

            time.sleep(5)

    except KeyboardInterrupt:
        pass