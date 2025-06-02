import time
import traceback
from pathlib import Path
from art import art
from openpyxl import load_workbook

from SupportFunctions import *


TOPIC = "save_measures_taken_in_db"
WORKER_ID = "python-worker-5"



def save_measures(data: dict, business_key):
    """
    Append or set the 'measures taken' field in the Excel database for the specified feedback entry,
    and update its status to 'review-board'.

    Args:
        data: Dictionary containing the documented measures.
        business_key: Unique identifier of the feedback entry to update.
    """

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # get row with the business key
    row_number = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[0]  # business key column
        if int(cell.value) == int(business_key):
            row_number = cell.row
            break


    if ws.cell(row_number, 17).value:
        measures_taken = ws.cell(row_number, 17).value
        ws.cell(row_number,17, f"{measures_taken}\n\n{data['measuresTaken']}")

    else:
        ws.cell(row_number, 17, data['measuresTaken'])

    ws.cell(row_number, 16, "review-board")

    # save excel
    wb.save(EXCEL_FILE)
    print("Data written to Excel.")



if __name__ == "__main__":
    print(f"Worker \"{Path(__file__).name}\" started — polling Camunda...")
    try:
        time.sleep(2)
        while True:
            try:
                for task in fetch_and_lock(worker_id=WORKER_ID, topic=TOPIC):
                    business_key = task.get("businessKey", "")
                    task_id = task["id"]
                    variables = {k: v["value"] for k, v in task["variables"].items()}
                    print(f"Worker \"{Path(__file__).name} fetched task {task_id} with business key {business_key}")
                    try:
                        save_measures(data=variables, business_key=business_key)
                        complete_task(task_id=task_id, variables=variables, worker_id=WORKER_ID)
                        print(f"Worker \"{Path(__file__).name} completed task {task_id} with business key {business_key}")
                    except PermissionError:
                        print("Excel in use, try again later")
                        time.sleep(30)
                    except Exception as exc:
                        print(f"Error in task {task_id}: {exc} {art('confused scratch')}")
                        traceback.print_exc()

            except Exception as exc:
                print(f"Fetch error: {exc} {art('table flip2')}")
                traceback.print_exc()

            time.sleep(5)

    except KeyboardInterrupt:
        pass