import requests
import time
from openpyxl import Workbook, load_workbook
from os.path import exists

CAMUNDA_ENGINE_URL = "https://digibp.engine.martinlab.science/engine-rest"
TOPIC = "store-feedback-in-db"
WORKER_ID = "python-worker-1"
EXCEL_FILE = "form_data.xlsx"


def fetch_and_lock():
    response = requests.post(f"{CAMUNDA_ENGINE_URL}/external-task/fetchAndLock", json={
        "workerId": WORKER_ID,
        "maxTasks": 1,
        "usePriority": False,
        "topics": [{
            "topicName": TOPIC,
            "lockDuration": 10000,
            "tenantId": "25DIGIBP12"
        }]
    })
    return response.json()


def complete_task(task_id, variables):
    requests.post(f"{CAMUNDA_ENGINE_URL}/external-task/{task_id}/complete", json={
        "workerId": WORKER_ID,
        "variables": {}
    })


def write_to_excel(data: dict):
    if exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, "")

    ws.append(list(data.values()))
    wb.save(EXCEL_FILE)
    print("Data written to Excel.")


while True:
    tasks = fetch_and_lock()
    for task in tasks:
        task_id = task['id']
        variables = {k: v['value'] for k, v in task['variables'].items()}
        print(f"Fetched task {task_id} with variables {variables}")

        write_to_excel(variables)
        complete_task(task_id, variables)
    time.sleep(5)