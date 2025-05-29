import time
import traceback
from pathlib import Path
from art import art
from openpyxl import load_workbook

from SupportFunctions import *

TOPIC = "save_classification_in_db"
WORKER_ID = "python-worker-4"



def save_classification(data: dict, business_key):

    wb = load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb.active

    # get row with the business key
    row_number = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[0]  # business key column
        if int(cell.value) == int(business_key):
            row_number = cell.row
            break

    if "feedbackText" in data:
        ws.cell(row_number, 9, data["feedbackText"])
    else:
        ws.cell(row_number, 9, "NA")

    if "feedbackType" in data:
        ws.cell(row_number, 3, data["feedbackType"])
    else:
        ws.cell(row_number, 3, "NA")

    if "urgency" in data:
        ws.cell(row_number, 11, data["urgency"])
    else:
        ws.cell(row_number, 11, "NA")

    if "impactScope" in data:
        ws.cell(row_number, 12, data["impactScope"])
    else:
        ws.cell(row_number, 12, "NA")

    if "forwardToDepartment" in data:
        ws.cell(row_number, 13, data["forwardToDepartment"])
    else:
        ws.cell(row_number, 13, "NA")


    # set status "open"
    ws.cell(row_number, 4, "")
    ws.cell(row_number, 16, "open")


    # save excel
    wb.save(EXCEL_FILE)
    print("Data written to Excel.")



if __name__ == "__main__":
    print(f"Worker \"{Path(__file__).name}\" started — polling Camunda...")
    try:
        time.sleep(2)
        while True:
            try:
                for task in fetch_and_lock(worker_id=WORKER_ID, topic=TOPIC):
                    business_key = task.get("businessKey", "")
                    task_id = task["id"]
                    variables = {k: v["value"] for k, v in task["variables"].items()}
                    print(f"Worker \"{Path(__file__).name} fetched task {task_id} with business key {business_key}")
                    try:
                        save_classification(data=variables, business_key=business_key)
                        complete_task(task_id=task_id, variables=variables, worker_id=WORKER_ID, send_variables="new")
                        print(f"Worker \"{Path(__file__).name} completed task {task_id} with business key {business_key}")
                    except Exception as exc:
                        print(f"Error in task {task_id}: {exc} {art('confused scratch')}")
                        traceback.print_exc()

            except Exception as exc:
                print(f"Fetch error: {exc} {art('table flip2')}")
                traceback.print_exc()

            time.sleep(5)

    except KeyboardInterrupt:
        pass