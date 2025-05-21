import requests
import time
import requests, time, pathlib, sqlite3
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from os.path import exists

CAMUNDA_ENGINE_URL = "https://digibp.engine.martinlab.science/engine-rest"
TOPIC = "store-feedback-in-db"
WORKER_ID = "python-worker-1"
EXCEL_FILE = "form_data.xlsx"

# openpyxl formating
bold = Font(bold=True)


def fetch_and_lock():
    response = requests.post(f"{CAMUNDA_ENGINE_URL}/external-task/fetchAndLock", json={
        "workerId": WORKER_ID,
        "maxTasks": 1,
        "usePriority": False,
        "topics": [{
            "topicName": TOPIC,
            "lockDuration": 10000,
            "tenantId": "25DIGIBP12"
        }]
    })
    return response.json()


def complete_task(task_id, variables):
    requests.post(f"{CAMUNDA_ENGINE_URL}/external-task/{task_id}/complete", json={
        "workerId": WORKER_ID,
        "variables": {}
    })


def write_to_excel(data: dict, task_id):
    if exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

        ws.cell(1, 1, "submissionID").font = bold
        ws.cell(1, 2, "feedbackDate").font = bold
        ws.cell(1, 3, "feedbackType").font = bold
        ws.cell(1, 4, "query").font = bold
        ws.cell(1, 5, "email").font = bold
        ws.cell(1, 6, "phone").font = bold
        ws.cell(1, 7, "firstName").font = bold
        ws.cell(1, 8, "lastName").font = bold
        ws.cell(1, 9, "feedbackText").font = bold
        ws.cell(1, 10, "needsClarification").font = bold
        ws.cell(1, 11, "urgency").font = bold
        ws.cell(1, 12, "impactScope").font = bold
        ws.cell(1, 13, "forwardToDepartment").font = bold
        ws.cell(1, 14, "linkToAdditForm").font = bold
        ws.cell(1, 15, "reminderSent").font = bold
        ws.cell(1, 16, "status").font = bold


    # get next empty row
    row = ws.max_row + 1

    print(data)

    feedback_date = str(date.today().strftime("%d.%m.%Y"))

    # write new data to excel
    ws.cell(row, 1, task_id)
    ws.cell(row, 2, feedback_date)
    ws.cell(row, 5, data["email"])
    ws.cell(row, 6, data["phone"])
    ws.cell(row, 7, data["firstName"])
    ws.cell(row, 8, data["lastName"])
    ws.cell(row,9,data["feedbackText"])


    # save excel
    wb.save(EXCEL_FILE)
    print("Data written to Excel.")

def store_in_db(data: dict, task_id: str) -> None:
    with sqlite3.connect(DB_FILE) as con:
        con.execute(
            """
            INSERT INTO feedback (
              camunda_task_id, created_at, feedback_date,
              feedback_type, query, email, phone,
              first_name, last_name, feedback_text,
              needs_clarification, urgency, impact_scope,
              forward_to_dept, link_to_add_form,
              reminder_sent, status
            )
            VALUES (
              :task_id, :created_at, :feedback_date,
              :feedback_type, :query, :email, :phone,
              :first_name, :last_name, :feedback_text,
              :needs_clarification, :urgency, :impact_scope,
              :forward_to_dept, :link_to_add_form,
              :reminder_sent, :status
            )
            """,
            {
                "task_id":            task_id,
                "created_at":         datetime.utcnow().isoformat(timespec="seconds"),
                "feedback_date":      date.today().strftime("%d.%m.%Y"),
                "feedback_type":      data.get("feedbackType"),
                "query":              data.get("query"),
                "email":              data.get("email"),
                "phone":              data.get("phone"),
                "first_name":         data.get("firstName"),
                "last_name":          data.get("lastName"),
                "feedback_text":      data.get("feedbackText"),
                "needs_clarification": data.get("needsClarification"),
                "urgency":            data.get("urgency"),
                "impact_scope":       data.get("impactScope"),
                "forward_to_dept":    data.get("forwardToDepartment"),
                "link_to_add_form":   data.get("linkToAdditForm"),
                "reminder_sent":      data.get("reminderSent"),
                "status":             data.get("status"),
            },
        )
    print("Row written to SQLite →", DB_FILE)




while True:
    try:
        for task in fetch_and_lock():
            task_id = task["id"]
            vars_ = {k: v["value"] for k, v in task["variables"].items()}
            print("Fetched", task_id)

            try:
                store_in_db(vars_, task_id)
                write_to_excel(vars_, task_id)
                complete_task(task_id)
            except Exception as ex:
                print("Error in task", task_id, ":", ex)

    except Exception as ex:
        print("Fetch/Lock error:", ex)