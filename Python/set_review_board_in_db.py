import requests
import time
import traceback
from art import tprint, art
from openpyxl import load_workbook



CAMUNDA_ENGINE_URL = "https://digibp.engine.martinlab.science/engine-rest"
TOPIC = "set_review_board"
WORKER_ID = "python-worker-644"
EXCEL_FILE = "form_data.xlsx"


def fetch_and_lock():
    response = requests.post(f"{CAMUNDA_ENGINE_URL}/external-task/fetchAndLock", json={
        "workerId": WORKER_ID,
        "maxTasks": 1,
        "usePriority": False,
        "topics": [{
            "topicName": TOPIC,
            "lockDuration": 10000,
            "tenantId": "25DIGIBP12"
        }]
    })
    return response.json()


def complete_task(task_id, variables):
    requests.post(f"{CAMUNDA_ENGINE_URL}/external-task/{task_id}/complete", json={
        "workerId": WORKER_ID,
        "variables": {}
    })


def set_status(business_key):

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # get row with the business key
    row_number = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[0]  # business key column
        if int(cell.value) == business_key:
            row_number = cell.row
            break

    ws.cell(row_number,16, "review-board")

    # save excel
    wb.save(EXCEL_FILE)
    print("Data written to Excel.")



if __name__ == "__main__":
   tprint("25-DIGIBP-1", font="small")
   print("Worker started — polling Camunda...")
   print(f"{art('hugger')}\n")
   while True:
       try:
           for task in fetch_and_lock():
               task_id = task["id"]
               variables = {k: v["value"] for k, v in task["variables"].items()}
               print(f"Fetched task {task_id}")
               try:
                   business_key = task.get("businessKey", "")
                   set_status(business_key)
                   complete_task(task_id, variables)
               except Exception as exc:
                   print(f"Error in task {task_id}: {exc} {art('confused scratch')}")
                   traceback.print_exc()
       except Exception as exc:
           print(f"Fetch error: {exc} {art('confused scratch')}")
           traceback.print_exc()

       time.sleep(5)