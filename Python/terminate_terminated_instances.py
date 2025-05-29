import time
import traceback
from pathlib import Path
from openpyxl import load_workbook
import requests

from SupportFunctions import *


# read config file
with open("config.json", "r") as f:
    config = json.load(f)
f.close()


CAMUNDA_MESSAGE_URL = f"{config['camundaEngineUrl']}/message"
TENANT_ID  = config["tenantID"]
MESSAGE    = "PROCESS_TERMINATED"
POLL_SEC   = 30
STATUS_COL = 16 # column P

def scan_and_terminate() -> None:
    db_path = Path(EXCEL_FILE)

    # exit quietly if the file does not exist yet
    if not db_path.exists():
        print("No feedback database yet – nothing to scan.")
        return

    # open and process
    wb = load_workbook(db_path, keep_vba=True)
    ws = wb.active
    rows_changed = False

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # skip header
        business_key = row[0].value  # column A
        status = row[STATUS_COL - 1].value  # column P

        if status == "terminate":
            payload = {
                "messageName": MESSAGE,
                "tenantId": TENANT_ID,
                "businessKey": str(business_key)
            }
            try:
                r = requests.post(CAMUNDA_MESSAGE_URL, json=payload, timeout=10)
                r.raise_for_status()
                print(f"Terminated process {business_key}")
                ws.cell(row=row[0].row, column=STATUS_COL, value="cancelled")
                rows_changed = True
            except Exception as exc:
                print(f"Could not terminate {business_key}: {exc}")

    if rows_changed:
        wb.save(db_path)
        print("Excel updated.\n")



if __name__ == "__main__":
    print(f"Worker \"{Path(__file__).name}\" started — watching for terminate requests...")
    try:
        time.sleep(2)
        while True:
            try:
                scan_and_terminate()
            except Exception:
                traceback.print_exc()

            time.sleep(POLL_SEC)

    except KeyboardInterrupt:
        pass
